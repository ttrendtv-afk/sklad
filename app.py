from flask import Flask, render_template, request, redirect, url_for, session, jsonify, flash, Response
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date
from functools import wraps
import os, json, csv, io, re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'wms-secret-2024-change-me')

# ─── DATABASE: supports both SQLite and PostgreSQL ─────────────────────────────
db_url = os.environ.get('DATABASE_URL', 'sqlite:///wms.db')
if db_url.startswith('postgres://'):
    db_url = db_url.replace('postgres://', 'postgresql://', 1)
app.config['SQLALCHEMY_DATABASE_URI'] = db_url
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False
app.config['SQLALCHEMY_ENGINE_OPTIONS'] = {'pool_pre_ping': True}
db = SQLAlchemy(app)

@app.after_request
def security_headers(response):
    response.headers['X-Content-Type-Options'] = 'nosniff'
    response.headers['X-Frame-Options'] = 'DENY'
    response.headers['X-XSS-Protection'] = '1; mode=block'
    return response

# ─── MODELS ───────────────────────────────────────────────────────────────────

class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), default='worker')
    name = db.Column(db.String(100))

class Store(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(100), unique=True, nullable=False)

class PhoneModel(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(150), unique=True, nullable=False)

class Phone(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    model_id = db.Column(db.Integer, db.ForeignKey('phone_model.id'), nullable=False)
    store_id = db.Column(db.Integer, db.ForeignKey('store.id'), nullable=False)
    imei = db.Column(db.String(20), unique=True, nullable=False)
    state = db.Column(db.String(20), default='НОВАЯ')
    cost_usd = db.Column(db.Float, default=0)
    markup_pct = db.Column(db.Float, default=20)
    gel_rate = db.Column(db.Float, default=2.7)
    price_gel = db.Column(db.Float, default=0)
    sold_price_gel = db.Column(db.Float)
    date_in = db.Column(db.Date, default=date.today)
    date_sold = db.Column(db.Date)
    date_returned = db.Column(db.Date)
    notes = db.Column(db.String(300))
    status = db.Column(db.String(20), default='на складе')
    added_by = db.Column(db.Integer, db.ForeignKey('user.id'))
    buyer = db.Column(db.String(100))
    model = db.relationship('PhoneModel', backref='phones')
    store = db.relationship('Store', backref='phones')

class ChangeLog(db.Model):
    """История всех изменений"""
    id = db.Column(db.Integer, primary_key=True)
    ts = db.Column(db.DateTime, default=datetime.utcnow)
    user_id = db.Column(db.Integer, db.ForeignKey('user.id'))
    action = db.Column(db.String(30))   # add / edit / sell / return / delete
    imei = db.Column(db.String(20))
    model_name = db.Column(db.String(150))
    detail = db.Column(db.String(300))
    user = db.relationship('User')

class Setting(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    key = db.Column(db.String(50), unique=True)
    value = db.Column(db.String(500))

class PhoneState(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(50), unique=True, nullable=False)
    sort = db.Column(db.Integer, default=0)

# ─── ROLES ────────────────────────────────────────────────────────────────────
ROLE_WEIGHTS = {'admin': 4, 'manager': 3, 'editor': 2, 'worker': 1}

def role_required(min_role):
    def decorator(f):
        @wraps(f)
        def dec(*a, **kw):
            if 'user_id' not in session:
                return redirect(url_for('login'))
            if ROLE_WEIGHTS.get(session.get('role'), 0) < ROLE_WEIGHTS.get(min_role, 99):
                flash('Нет доступа', 'error')
                return redirect(url_for('catalog'))
            return f(*a, **kw)
        return dec
    return decorator

login_required = role_required('worker')

# ─── HELPERS ──────────────────────────────────────────────────────────────────

def get_setting(key, default=''):
    s = Setting.query.filter_by(key=key).first()
    return s.value if s else default

def set_setting(key, value):
    s = Setting.query.filter_by(key=key).first()
    if s: s.value = str(value)
    else: db.session.add(Setting(key=key, value=str(value)))
    db.session.commit()

def get_states():
    states = [s.name for s in PhoneState.query.order_by(PhoneState.sort).all()]
    return states or ['НОВАЯ', 'REF.', 'Б/У']

def log_action(action, phone, detail=''):
    try:
        entry = ChangeLog(
            user_id=session.get('user_id'),
            action=action,
            imei=phone.imei if phone else '',
            model_name=f"{phone.model.name}" if phone else '',
            detail=detail
        )
        db.session.add(entry)
    except: pass

def send_telegram(message):
    pass  # Telegram disabled

def phone_to_dict(p):
    gel_rate = p.gel_rate or float(get_setting('gel_rate', '2.7'))
    eff = p.sold_price_gel if p.sold_price_gel is not None else p.price_gel
    profit_gel = round((eff or 0) - (p.cost_usd or 0) * gel_rate, 2)
    return {
        'id': p.id, 'imei': p.imei, 'model': p.model.name, 'model_id': p.model_id,
        'store': p.store.name, 'store_id': p.store_id, 'state': p.state,
        'cost_usd': p.cost_usd or 0, 'markup_pct': p.markup_pct or 20,
        'gel_rate': gel_rate, 'price_gel': p.price_gel or 0,
        'sold_price_gel': p.sold_price_gel,
        'date_in': str(p.date_in) if p.date_in else '',
        'date_sold': str(p.date_sold) if p.date_sold else '',
        'date_returned': str(p.date_returned) if p.date_returned else '',
        'notes': p.notes or '', 'status': p.status, 'buyer': p.buyer or '',
        'profit_gel': profit_gel,
    }

# ─── AUTH ─────────────────────────────────────────────────────────────────────

@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'POST':
        u = User.query.filter_by(username=request.form['username']).first()
        if u and check_password_hash(u.password_hash, request.form['password']):
            session.update({'user_id': u.id, 'username': u.username,
                            'role': u.role, 'name': u.name or u.username})
            return redirect(url_for('catalog'))
        flash('Неверный логин или пароль', 'error')
    return render_template('login.html')

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

@app.route('/change-password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        current = request.form.get('current_password', '')
        new_pw = request.form.get('new_password', '')
        confirm = request.form.get('confirm_password', '')
        user = User.query.get(session['user_id'])
        if not check_password_hash(user.password_hash, current):
            flash('Неверный текущий пароль', 'error')
            return render_template('change_password.html')
        if len(new_pw) < 4:
            flash('Пароль должен быть не менее 4 символов', 'error')
            return render_template('change_password.html')
        if new_pw != confirm:
            flash('Пароли не совпадают', 'error')
            return render_template('change_password.html')
        user.password_hash = generate_password_hash(new_pw)
        db.session.commit()
        flash('Пароль успешно изменён', 'success')
        return redirect(url_for('catalog'))
    return render_template('change_password.html')

@app.route('/admin/reset-password/<int:uid>', methods=['POST'])
@role_required('admin')
def reset_password(uid):
    new_pw = request.form.get('new_password', '')
    if len(new_pw) < 4:
        flash('Пароль должен быть не менее 4 символов', 'error')
        return redirect(url_for('references'))
    u = User.query.get_or_404(uid)
    u.password_hash = generate_password_hash(new_pw)
    db.session.commit()
    flash(f'Пароль пользователя {u.username} изменён', 'success')
    return redirect(url_for('references'))

# ─── CATALOG ──────────────────────────────────────────────────────────────────

@app.route('/')
@login_required
def catalog():
    models = PhoneModel.query.order_by(PhoneModel.name).all()
    stores = Store.query.order_by(Store.name).all()
    gel_rate = get_setting('gel_rate', '2.7')
    states = get_states()
    # Low stock warning
    low_stock = []
    for m in models:
        cnt = Phone.query.filter_by(model_id=m.id, status='на складе').count()
        if cnt <= 1:
            low_stock.append({'model': m.name, 'count': cnt})
    return render_template('catalog.html', models=models, stores=stores,
                           gel_rate=gel_rate, states=states, low_stock=low_stock)

@app.route('/api/catalog')
@login_required
def api_catalog():
    sel_imei = request.args.get('imei', '')
    sel_model = request.args.get('model', '')
    sel_store = request.args.get('store', '')
    sel_state = request.args.get('state', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    q = Phone.query.filter_by(status='на складе')
    if sel_imei: q = q.filter(Phone.imei.ilike(f'%{sel_imei}%'))
    if sel_model: q = q.filter(Phone.model_id == sel_model)
    if sel_store: q = q.filter(Phone.store_id == sel_store)
    if sel_state: q = q.filter(Phone.state == sel_state)
    if date_from:
        try: q = q.filter(Phone.date_in >= date.fromisoformat(date_from))
        except: pass
    if date_to:
        try: q = q.filter(Phone.date_in <= date.fromisoformat(date_to))
        except: pass

    phones = q.order_by(Phone.model_id, Phone.store_id).all()
    from collections import defaultdict
    groups = defaultdict(lambda: {'model_id': 0, 'model': '', 'count': 0, 'states': {}, 'phones': []})
    for p in phones:
        k = p.model.name
        groups[k]['model_id'] = p.model_id
        groups[k]['model'] = k
        groups[k]['count'] += 1
        groups[k]['states'][p.state] = groups[k]['states'].get(p.state, 0) + 1
        groups[k]['phones'].append(phone_to_dict(p))
    return jsonify(list(groups.values()))

@app.route('/api/model-phones/<int:model_id>')
@login_required
def api_model_phones(model_id):
    sel_store = request.args.get('store', '')
    sel_state = request.args.get('state', '')
    sel_imei = request.args.get('imei', '')
    q = Phone.query.filter_by(model_id=model_id, status='на складе')
    if sel_store: q = q.filter(Phone.store_id == sel_store)
    if sel_state: q = q.filter(Phone.state == sel_state)
    if sel_imei: q = q.filter(Phone.imei.ilike(f'%{sel_imei}%'))
    phones = q.all()
    stores = Store.query.all()
    return jsonify({'phones': [phone_to_dict(p) for p in phones],
                    'stores': [{'id': s.id, 'name': s.name} for s in stores]})

@app.route('/api/phone/<imei>')
@login_required
def api_phone(imei):
    p = Phone.query.filter_by(imei=imei).first_or_404()
    stores = Store.query.all()
    states = get_states()
    d = phone_to_dict(p)
    d['stores'] = [{'id': s.id, 'name': s.name} for s in stores]
    d['states'] = states
    return jsonify(d)

@app.route('/phone/<imei>/edit', methods=['POST'])
@role_required('editor')
def phone_edit(imei):
    p = Phone.query.filter_by(imei=imei).first_or_404()
    old_status = p.status
    old_store = p.store.name

    p.state = request.form.get('state', p.state)
    p.store_id = int(request.form.get('store_id', p.store_id))
    p.cost_usd = float(request.form.get('cost_usd') or 0)
    p.markup_pct = float(request.form.get('markup_pct') or 20)
    p.gel_rate = float(request.form.get('gel_rate') or 2.7)
    p.price_gel = float(request.form.get('price_gel') or 0)
    sp = request.form.get('sold_price_gel')
    p.sold_price_gel = float(sp) if sp else None
    p.notes = request.form.get('notes', '')
    p.buyer = request.form.get('buyer', '')

    di = request.form.get('date_in')
    if di:
        try: p.date_in = date.fromisoformat(di)
        except: pass
    ds = request.form.get('date_sold')
    p.date_sold = date.fromisoformat(ds) if ds else None
    dr = request.form.get('date_returned')
    p.date_returned = date.fromisoformat(dr) if dr else None

    if p.date_returned: p.status = 'возврат'
    elif p.date_sold: p.status = 'продан'
    else: p.status = 'на складе'

    # Log the action
    action = 'edit'
    detail_parts = []
    if p.status == 'продан' and old_status != 'продан':
        action = 'sell'
        detail_parts.append(f'Продан за {p.sold_price_gel or p.price_gel} ₾')
        if p.buyer: detail_parts.append(f'Покупатель: {p.buyer}')
        send_telegram(
            f'💸 <b>Продажа</b>\n'
            f'📱 {p.model.name}\n'
            f'IMEI: {p.imei}\n'
            f'Магазин: {p.store.name}\n'
            f'Цена: {p.sold_price_gel or p.price_gel} ₾\n'
            f'Продавец: {session.get("name", "—")}'
        )
        # Check remaining stock after sale
        remaining = Phone.query.filter_by(model_id=p.model_id, status='на складе').count() - 1
        if remaining <= 0:
            send_telegram(f'⚠️ <b>Заканчивается товар!</b>\n📱 {p.model.name} — остаток: {remaining} шт')
    elif p.status == 'возврат' and old_status != 'возврат':
        action = 'return'
        detail_parts.append(f'Возврат из {old_store}')
    elif old_store != p.store.name:
        action = 'move'
        detail_parts.append(f'Перемещён: {old_store} → {p.store.name}')
    else:
        detail_parts.append('Редактирование данных')

    log_action(action, p, '; '.join(detail_parts))
    db.session.commit()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True})
    flash('Сохранено', 'success')
    return redirect(url_for('catalog'))

@app.route('/phone/<imei>/delete', methods=['POST'])
@role_required('admin')
def phone_delete(imei):
    # Double confirmation check
    confirm = request.form.get('confirm_delete', '') or request.headers.get('X-Confirm-Delete', '')
    if confirm != 'yes':
        return jsonify({'ok': False, 'error': 'Требуется подтверждение'}), 400

    p = Phone.query.filter_by(imei=imei).first_or_404()
    log_action('delete', p, 'Товар удалён')
    db.session.commit()
    db.session.delete(p)
    db.session.commit()

    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
        return jsonify({'ok': True})
    flash('Удалено', 'success')
    return redirect(url_for('catalog'))

# ─── IMPORT ───────────────────────────────────────────────────────────────────

@app.route('/import', methods=['GET', 'POST'])
@role_required('editor')
def import_phones():
    models = PhoneModel.query.order_by(PhoneModel.name).all()
    stores = Store.query.order_by(Store.name).all()
    gel_rate = get_setting('gel_rate', '2.7')
    states = get_states()

    if request.method == 'POST':
        f = request.files.get('file')
        if not f:
            flash('Файл не выбран', 'error')
            return redirect(url_for('import_phones'))

        model_id = int(request.form['model_id'])
        store_id = int(request.form['store_id'])
        state = request.form.get('state', states[0] if states else 'НОВАЯ')
        cost_usd = float(request.form.get('cost_usd') or 0)
        markup_pct = float(request.form.get('markup_pct') or 20)
        gel_rate_val = float(request.form.get('gel_rate') or 2.7)
        date_in_str = request.form.get('date_in') or str(date.today())
        manual_price = request.form.get('price_gel', '').strip()
        if manual_price:
            price_gel = float(manual_price)
        elif cost_usd:
            price_gel = round(cost_usd * (1 + markup_pct / 100) * gel_rate_val, 2)
        else:
            price_gel = 0

        raw = f.read()
        try:
            content = raw.decode('utf-8-sig', errors='replace')
        except Exception:
            content = raw.decode('latin-1', errors='replace')

        all_imeis = re.findall(r'\b\d{10,16}\b', content)
        added, skipped = 0, 0
        m_obj = PhoneModel.query.get(model_id)
        s_obj = Store.query.get(store_id)

        for imei in all_imeis:
            imei = imei.strip()
            if not imei or len(imei) < 10:
                continue
            if Phone.query.filter_by(imei=imei).first():
                skipped += 1
                continue
            p = Phone(
                model_id=model_id, store_id=store_id, imei=imei,
                state=state, cost_usd=cost_usd, markup_pct=markup_pct,
                gel_rate=gel_rate_val, price_gel=price_gel,
                date_in=date.fromisoformat(date_in_str),
                added_by=session['user_id'],
            )
            db.session.add(p)
            db.session.flush()
            log_action('add', p, f'Импорт партии. Цена: {price_gel} ₾')
            added += 1

        db.session.commit()

        if added > 0 and m_obj and s_obj:
            send_telegram(
                f'📦 <b>Поступление товара</b>\n'
                f'📱 {m_obj.name}\n'
                f'Количество: {added} шт\n'
                f'Магазин: {s_obj.name}\n'
                f'Добавил: {session.get("name", "—")}'
            )

        flash(f'Добавлено: {added} шт. Пропущено (дубли): {skipped} шт.', 'success')
        return redirect(url_for('catalog'))

    return render_template('import.html', models=models, stores=stores,
                           gel_rate=gel_rate, today=str(date.today()), states=states)

@app.route('/import/template')
@login_required
def import_template():
    content = "imei\n351111111111111\n351111111111222\n"
    return Response(content, mimetype='text/csv',
                    headers={'Content-Disposition': 'attachment;filename=imei_template.csv'})

# ─── FINANCE ──────────────────────────────────────────────────────────────────

@app.route('/finance')
@role_required('manager')
def finance():
    models = PhoneModel.query.order_by(PhoneModel.name).all()
    stores = Store.query.order_by(Store.name).all()
    gel_rate = float(get_setting('gel_rate', '2.7'))
    states = get_states()
    return render_template('finance.html', models=models, stores=stores,
                           gel_rate=gel_rate, states=states)

@app.route('/api/finance')
@role_required('manager')
def api_finance():
    sel_store = request.args.get('store', '')
    sel_model = request.args.get('model', '')
    sel_state = request.args.get('state', '')
    sel_imei = request.args.get('imei', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    gel_rate = float(get_setting('gel_rate', '2.7'))

    q = Phone.query.filter_by(status='на складе')
    if sel_store: q = q.filter(Phone.store_id == sel_store)
    if sel_model: q = q.filter(Phone.model_id == sel_model)
    if sel_state: q = q.filter(Phone.state == sel_state)
    if sel_imei: q = q.filter(Phone.imei.ilike(f'%{sel_imei}%'))
    if date_from:
        try: q = q.filter(Phone.date_in >= date.fromisoformat(date_from))
        except: pass
    if date_to:
        try: q = q.filter(Phone.date_in <= date.fromisoformat(date_to))
        except: pass

    phones = q.order_by(Phone.model_id).all()
    total_price = total_cost = 0
    result = []
    for p in phones:
        gr = p.gel_rate or gel_rate
        total_price += p.price_gel or 0
        total_cost += (p.cost_usd or 0) * gr
        result.append(phone_to_dict(p))

    return jsonify({'phones': result, 'count': len(result),
                    'total_price': round(total_price, 2),
                    'total_cost': round(total_cost, 2),
                    'total_profit': round(total_price - total_cost, 2)})

@app.route('/finance/bulk', methods=['POST'])
@role_required('manager')
def finance_bulk():
    markup = float(request.form.get('markup', 20))
    gel_rate = float(request.form.get('gel_rate', 2.7))
    sel_store = request.form.get('store', '')
    sel_model = request.form.get('model', '')
    sel_state = request.form.get('state', '')

    q = Phone.query.filter_by(status='на складе')
    if sel_store: q = q.filter(Phone.store_id == sel_store)
    if sel_model: q = q.filter(Phone.model_id == sel_model)
    if sel_state: q = q.filter(Phone.state == sel_state)

    phones = q.all()
    for p in phones:
        if p.cost_usd:
            p.markup_pct = markup
            p.gel_rate = gel_rate
            p.price_gel = round(p.cost_usd * (1 + markup / 100) * gel_rate, 2)
    set_setting('gel_rate', str(gel_rate))
    db.session.commit()
    flash(f'Обновлено {len(phones)} позиций', 'success')
    return redirect(url_for('finance'))

# ─── REPORTS ──────────────────────────────────────────────────────────────────

@app.route('/reports')
@role_required('manager')
def reports():
    tab = request.args.get('tab', 'sales')
    if tab not in ['sales', 'returns']:
        tab = 'sales'

    models = PhoneModel.query.order_by(PhoneModel.name).all()
    stores = Store.query.order_by(Store.name).all()
    states = get_states()
    gel_rate = float(get_setting('gel_rate', '2.7'))

    return render_template('reports.html', tab=tab, models=models, stores=stores,
                           states=states, gel_rate=gel_rate)

@app.route('/api/reports')
@role_required('manager')
def api_reports():
    tab = request.args.get('tab', 'sales')
    sel_store = request.args.get('store', '')
    sel_model = request.args.get('model', '')
    sel_state = request.args.get('state', '')
    sel_imei = request.args.get('imei', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    gel_rate = float(get_setting('gel_rate', '2.7'))

    status = 'продан' if tab == 'sales' else 'возврат'
    q = Phone.query.filter_by(status=status)
    if sel_store: q = q.filter(Phone.store_id == sel_store)
    if sel_model: q = q.filter(Phone.model_id == sel_model)
    if sel_state: q = q.filter(Phone.state == sel_state)
    if sel_imei: q = q.filter(Phone.imei.ilike(f'%{sel_imei}%'))
    date_field = Phone.date_sold if tab == 'sales' else Phone.date_returned
    if date_from:
        try: q = q.filter(date_field >= date.fromisoformat(date_from))
        except: pass
    if date_to:
        try: q = q.filter(date_field <= date.fromisoformat(date_to))
        except: pass
    phones = q.order_by(date_field.desc()).all()
    return jsonify({'phones': [phone_to_dict(p) for p in phones]})

@app.route('/api/report-stats')
@role_required('manager')
def api_report_stats():
    sel_store = request.args.get('store', '')
    sel_model = request.args.get('model', '')
    sel_state = request.args.get('state', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    gel_rate = float(get_setting('gel_rate', '2.7'))

    def fq(status, df):
        q = Phone.query.filter_by(status=status)
        if sel_store: q = q.filter(Phone.store_id == sel_store)
        if sel_model: q = q.filter(Phone.model_id == sel_model)
        if sel_state: q = q.filter(Phone.state == sel_state)
        if date_from:
            try: q = q.filter(df >= date.fromisoformat(date_from))
            except: pass
        if date_to:
            try: q = q.filter(df <= date.fromisoformat(date_to))
            except: pass
        return q.all()

    sold = fq('продан', Phone.date_sold)
    returns = fq('возврат', Phone.date_returned)
    total_revenue = sum((p.sold_price_gel or p.price_gel or 0) for p in sold)
    total_cost = sum((p.cost_usd or 0) * (p.gel_rate or gel_rate) for p in sold)
    total_profit = total_revenue - total_cost
    total_returns = sum((p.sold_price_gel or p.price_gel or 0) for p in returns)
    net_gel = total_profit - total_returns

    return jsonify({
        'count_sold': len(sold), 'count_returns': len(returns),
        'total_revenue': round(total_revenue, 2),
        'total_profit': round(total_profit, 2),
        'total_returns': round(total_returns, 2),
        'net_gel': round(net_gel, 2),
    })

@app.route('/reports/export')
@role_required('manager')
def reports_export():
    tab = request.args.get('tab', 'sales')
    sel_store = request.args.get('store', '')
    sel_model = request.args.get('model', '')
    sel_state = request.args.get('state', '')
    sel_imei = request.args.get('imei', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')
    gel_rate = float(get_setting('gel_rate', '2.7'))

    status_map = {'sales': 'продан', 'returns': 'возврат', 'finance': 'на складе'}
    q = Phone.query.filter_by(status=status_map.get(tab, 'продан'))
    if sel_store: q = q.filter(Phone.store_id == sel_store)
    if sel_model: q = q.filter(Phone.model_id == sel_model)
    if sel_state: q = q.filter(Phone.state == sel_state)
    if sel_imei: q = q.filter(Phone.imei.ilike(f'%{sel_imei}%'))
    if tab == 'sales' and date_from:
        try: q = q.filter(Phone.date_sold >= date.fromisoformat(date_from))
        except: pass
    if tab == 'sales' and date_to:
        try: q = q.filter(Phone.date_sold <= date.fromisoformat(date_to))
        except: pass
    phones = q.all()

    wb = Workbook()
    ws = wb.active
    ws.title = tab
    hf = Font(bold=True, color='FFFFFF')
    hfill = PatternFill('solid', fgColor='2563EB')

    if tab == 'finance':
        headers = ['Модель','IMEI','Магазин','Состояние','Себест. $','Наценка %','Курс GEL','Цена ₾','Потенц. прибыль ₾']
    elif tab == 'sales':
        headers = ['Модель','IMEI','Магазин','Состояние','Себест. $','Курс GEL','Цена ₾','Цена продажи ₾','Прибыль ₾','Дата приёма','Дата продажи','Покупатель','Заметки']
    else:
        headers = ['Модель','IMEI','Магазин','Состояние','Цена ₾','Дата продажи','Дата возврата','Заметки']

    ws.append(headers)
    for cell in ws[1]:
        cell.font = hf
        cell.fill = hfill
        cell.alignment = Alignment(horizontal='center')

    for p in phones:
        gr = p.gel_rate or gel_rate
        eff = p.sold_price_gel if p.sold_price_gel is not None else p.price_gel
        profit = round((eff or 0) - (p.cost_usd or 0) * gr, 2)
        if tab == 'finance':
            pot = round((p.price_gel or 0) - (p.cost_usd or 0) * gr, 2)
            ws.append([p.model.name, p.imei, p.store.name, p.state,
                       p.cost_usd or 0, p.markup_pct or 0, gr, p.price_gel or 0, pot])
        elif tab == 'sales':
            ws.append([p.model.name, p.imei, p.store.name, p.state,
                       p.cost_usd or 0, gr, p.price_gel or 0, eff or 0, profit,
                       str(p.date_in or ''), str(p.date_sold or ''), p.buyer or '', p.notes or ''])
        else:
            ws.append([p.model.name, p.imei, p.store.name, p.state,
                       eff or 0, str(p.date_sold or ''), str(p.date_returned or ''), p.notes or ''])

    for col in ws.columns:
        w = max(len(str(c.value or '')) for c in col)
        ws.column_dimensions[col[0].column_letter].width = min(w + 4, 40)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return Response(buf.getvalue(),
                    mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                    headers={'Content-Disposition': f'attachment;filename=export_{tab}.xlsx'})

@app.route('/update-rate', methods=['POST'])
@role_required('manager')
def update_rate():
    gel = request.form.get('gel_rate')
    if gel: set_setting('gel_rate', gel)
    flash('Курс обновлён', 'success')
    return redirect(request.referrer or url_for('catalog'))

# ─── CHANGELOG ────────────────────────────────────────────────────────────────

@app.route('/changelog')
@role_required('manager')
def changelog():
    users = User.query.order_by(User.name).all()
    return render_template('changelog.html', users=users)

@app.route('/api/changelog')
@role_required('manager')
def api_changelog():
    page = request.args.get('page', 1, type=int)
    per_page = 50
    action = request.args.get('action', '')
    user_id = request.args.get('user_id', '')
    imei = request.args.get('imei', '')
    model = request.args.get('model', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    q = ChangeLog.query
    if action: q = q.filter(ChangeLog.action == action)
    if user_id: q = q.filter(ChangeLog.user_id == int(user_id))
    if imei: q = q.filter(ChangeLog.imei.ilike(f'%{imei}%'))
    if model: q = q.filter(ChangeLog.model_name.ilike(f'%{model}%'))
    if date_from:
        try: q = q.filter(ChangeLog.ts >= datetime.fromisoformat(date_from))
        except: pass
    if date_to:
        try: q = q.filter(ChangeLog.ts <= datetime.fromisoformat(date_to + ' 23:59:59'))
        except: pass

    total = q.count()
    pages = max(1, (total + per_page - 1) // per_page)
    items = q.order_by(ChangeLog.id.desc()).offset((page-1)*per_page).limit(per_page).all()

    return jsonify({
        'total': total, 'page': page, 'pages': pages,
        'items': [{
            'ts': log.ts.strftime('%d.%m.%Y %H:%M'),
            'user': log.user.name if log.user else '—',
            'action': log.action,
            'imei': log.imei or '',
            'model_name': log.model_name or '',
            'detail': log.detail or '',
        } for log in items]
    })

# ─── REFERENCES ───────────────────────────────────────────────────────────────

@app.route('/references')
@role_required('admin')
def references():
    models = PhoneModel.query.order_by(PhoneModel.name).all()
    stores = Store.query.order_by(Store.name).all()
    users = User.query.all()
    states = PhoneState.query.order_by(PhoneState.sort).all()
    return render_template('references.html', models=models, stores=stores, users=users, states=states)

@app.route('/references/model/add', methods=['POST'])
@role_required('admin')
def add_model():
    name = request.form['name'].strip()
    if name and not PhoneModel.query.filter_by(name=name).first():
        db.session.add(PhoneModel(name=name))
        db.session.commit()
        flash(f'Модель «{name}» добавлена', 'success')
    return redirect(url_for('references'))

@app.route('/references/model/edit/<int:mid>', methods=['POST'])
@role_required('admin')
def edit_model(mid):
    m = PhoneModel.query.get_or_404(mid)
    m.name = request.form['name'].strip()
    db.session.commit()
    flash('Модель обновлена', 'success')
    return redirect(url_for('references'))

@app.route('/references/model/delete/<int:mid>', methods=['POST'])
@role_required('admin')
def delete_model(mid):
    m = PhoneModel.query.get_or_404(mid)
    if m.phones:
        flash('Нельзя удалить — есть товары', 'error')
    else:
        db.session.delete(m)
        db.session.commit()
        flash('Модель удалена', 'success')
    return redirect(url_for('references'))

@app.route('/references/store/add', methods=['POST'])
@role_required('admin')
def add_store():
    name = request.form['name'].strip().upper()
    if name and not Store.query.filter_by(name=name).first():
        db.session.add(Store(name=name))
        db.session.commit()
        flash(f'Магазин «{name}» добавлен', 'success')
    return redirect(url_for('references'))

@app.route('/references/store/delete/<int:sid>', methods=['POST'])
@role_required('admin')
def delete_store(sid):
    s = Store.query.get_or_404(sid)
    if s.phones:
        flash('Нельзя удалить — есть товары', 'error')
    else:
        db.session.delete(s)
        db.session.commit()
        flash('Магазин удалён', 'success')
    return redirect(url_for('references'))

@app.route('/references/state/add', methods=['POST'])
@role_required('admin')
def add_state():
    name = request.form['name'].strip()
    if name and not PhoneState.query.filter_by(name=name).first():
        db.session.add(PhoneState(name=name, sort=PhoneState.query.count()))
        db.session.commit()
        flash(f'Состояние «{name}» добавлено', 'success')
    return redirect(url_for('references'))

@app.route('/references/state/delete/<int:sid>', methods=['POST'])
@role_required('admin')
def delete_state(sid):
    s = PhoneState.query.get_or_404(sid)
    db.session.delete(s)
    db.session.commit()
    flash('Состояние удалено', 'success')
    return redirect(url_for('references'))

@app.route('/references/user/add', methods=['POST'])
@role_required('admin')
def add_user():
    username = request.form['username'].strip()
    if User.query.filter_by(username=username).first():
        flash('Пользователь уже существует', 'error')
        return redirect(url_for('references'))
    db.session.add(User(username=username,
                        password_hash=generate_password_hash(request.form['password']),
                        name=request.form.get('name', ''),
                        role=request.form.get('role', 'worker')))
    db.session.commit()
    flash(f'Пользователь {username} создан', 'success')
    return redirect(url_for('references'))

@app.route('/references/user/delete/<int:uid>', methods=['POST'])
@role_required('admin')
def delete_user(uid):
    if uid == session['user_id']:
        flash('Нельзя удалить себя', 'error')
        return redirect(url_for('references'))
    u = User.query.get_or_404(uid)
    db.session.delete(u)
    db.session.commit()
    flash('Удалён', 'success')
    return redirect(url_for('references'))



# ─── BACKUP ───────────────────────────────────────────────────────────────────

@app.route('/backup/export')
@role_required('admin')
def backup_export():
    phones = Phone.query.all()
    data = [phone_to_dict(p) for p in phones]
    out = json.dumps({
        'phones': data,
        'models': [m.name for m in PhoneModel.query.all()],
        'stores': [s.name for s in Store.query.all()],
        'states': [s.name for s in PhoneState.query.order_by(PhoneState.sort).all()],
    }, ensure_ascii=False, indent=2)
    return Response(out, mimetype='application/json',
                    headers={'Content-Disposition': 'attachment;filename=backup_wms.json'})

@app.route('/backup/import', methods=['POST'])
@role_required('admin')
def backup_import():
    f = request.files.get('file')
    if not f:
        flash('Файл не выбран', 'error')
        return redirect(url_for('references'))
    try:
        data = json.load(f)
        for mname in data.get('models', []):
            if not PhoneModel.query.filter_by(name=mname).first():
                db.session.add(PhoneModel(name=mname))
        for sname in data.get('stores', []):
            if not Store.query.filter_by(name=sname).first():
                db.session.add(Store(name=sname))
        for stname in data.get('states', []):
            if not PhoneState.query.filter_by(name=stname).first():
                db.session.add(PhoneState(name=stname, sort=PhoneState.query.count()))
        db.session.flush()
        count = 0
        for row in data.get('phones', []):
            if Phone.query.filter_by(imei=row['imei']).first():
                continue
            m = PhoneModel.query.filter_by(name=row['model']).first()
            s = Store.query.filter_by(name=row['store']).first()
            if not m or not s: continue
            p = Phone(
                model_id=m.id, store_id=s.id, imei=row['imei'],
                state=row.get('state','НОВАЯ'), cost_usd=row.get('cost_usd',0),
                markup_pct=row.get('markup_pct',20), gel_rate=row.get('gel_rate',2.7),
                price_gel=row.get('price_gel',0), sold_price_gel=row.get('sold_price_gel'),
                date_in=date.fromisoformat(row['date_in']) if row.get('date_in') else None,
                date_sold=date.fromisoformat(row['date_sold']) if row.get('date_sold') else None,
                date_returned=date.fromisoformat(row['date_returned']) if row.get('date_returned') else None,
                status=row.get('status','на складе'), notes=row.get('notes',''), buyer=row.get('buyer',''),
            )
            db.session.add(p)
            count += 1
        db.session.commit()
        flash(f'Импортировано {count} позиций', 'success')
    except Exception as e:
        flash(f'Ошибка: {e}', 'error')
    return redirect(url_for('references'))


def _add_demo_phones():
    """Add demo phones for testing"""
    import random
    random.seed(99)
    stores = Store.query.all()
    models = PhoneModel.query.all()
    if not stores or not models:
        return
    states = ['НОВАЯ', 'REF.', 'Б/У', 'НОВАЯ', 'НОВАЯ']  # weight towards НОВАЯ
    gel_rate = 2.7
    used = set()

    def rand_imei():
        while True:
            imei = '3' + ''.join([str(random.randint(0,9)) for _ in range(14)])
            if imei not in used:
                used.add(imei)
                return imei

    demo_data = [
        # (model_idx, store_idx, state, cost_usd, status, date_sold)
        (0,  0, 'НОВАЯ',  380, 'на складе', None),
        (0,  1, 'REF.',   300, 'на складе', None),
        (0,  2, 'Б/У',    250, 'продан',    '2026-05-10'),
        (1,  0, 'НОВАЯ',  390, 'на складе', None),
        (2,  1, 'НОВАЯ',  470, 'на складе', None),
        (2,  2, 'НОВАЯ',  470, 'продан',    '2026-05-12'),
        (3,  0, 'REF.',   400, 'на складе', None),
        (4,  1, 'НОВАЯ',  480, 'на складе', None),
        (5,  2, 'НОВАЯ',  620, 'на складе', None),
        (5,  0, 'Б/У',    480, 'продан',    '2026-05-08'),
        (6,  1, 'НОВАЯ',  630, 'на складе', None),
        (7,  2, 'REF.',   520, 'на складе', None),
        (8,  0, 'НОВАЯ',  650, 'на складе', None),
        (8,  1, 'НОВАЯ',  650, 'продан',    '2026-05-14'),
        (9,  2, 'НОВАЯ',  780, 'на складе', None),
        (10, 0, 'REF.',   600, 'на складе', None),
        (11, 1, 'НОВАЯ',  800, 'на складе', None),
        (12, 2, 'НОВАЯ',  950, 'на складе', None),
        (13, 0, 'Б/У',    750, 'возврат',   '2026-05-13'),
        (14, 1, 'НОВАЯ', 1050, 'на складе', None),
        (15, 2, 'НОВАЯ', 1100, 'на складе', None),
        (16, 0, 'REF.',   850, 'на складе', None),
        (17, 1, 'НОВАЯ',  900, 'на складе', None),
        (18, 2, 'НОВАЯ',  960, 'продан',    '2026-05-15'),
        (19, 0, 'НОВАЯ', 1200, 'на складе', None),
    ]

    for m_i, s_i, state, cost, status, ds in demo_data:
        if m_i >= len(models) or s_i >= len(stores):
            continue
        markup = 20
        price = round(cost * (1 + markup/100) * gel_rate, 2)
        sold_price = round(price * random.uniform(0.95, 1.05), 2) if status == 'продан' else None
        p = Phone(
            model_id=models[m_i].id,
            store_id=stores[s_i % len(stores)].id,
            imei=rand_imei(),
            state=state,
            cost_usd=cost,
            markup_pct=markup,
            gel_rate=gel_rate,
            price_gel=price,
            sold_price_gel=sold_price,
            date_in=date(2026, 5, 1),
            date_sold=date.fromisoformat(ds) if ds else None,
            date_returned=date(2026, 5, 13) if status == 'возврат' else None,
            status=status,
            added_by=1,
        )
        db.session.add(p)
    db.session.commit()

# ─── INIT ─────────────────────────────────────────────────────────────────────

def init_db():
    db.create_all()
    if not User.query.filter_by(username='admin').first():
        db.session.add(User(username='admin', password_hash=generate_password_hash('admin123'),
                            role='admin', name='Администратор'))
    if not Store.query.filter_by(name='СКЛАД').first():
        for s in ['СКЛАД', 'МАГАЗИН 1', 'МАГАЗИН 2', 'МАГАЗИН 3']:
            db.session.add(Store(name=s))
    if not Setting.query.filter_by(key='gel_rate').first():
        db.session.add(Setting(key='gel_rate', value='2.7'))
    if not PhoneState.query.first():
        for i, s in enumerate(['НОВАЯ', 'REF.', 'Б/У', 'БРАК']):
            db.session.add(PhoneState(name=s, sort=i))
    if not Phone.query.first():
        _add_demo_phones()
    if not PhoneModel.query.first():
        for name in [
            'iPhone 13 128GB', 'iPhone 13 256GB',
            'iPhone 13 Pro 128GB', 'iPhone 13 Pro 256GB',
            'iPhone 14 128GB', 'iPhone 14 256GB',
            'iPhone 14 Pro 128GB', 'iPhone 14 Pro 256GB',
            'iPhone 14 Pro Max 256GB', 'iPhone 14 Pro Max 512GB',
            'iPhone 15 128GB', 'iPhone 15 256GB',
            'iPhone 15 Pro 128GB', 'iPhone 15 Pro 256GB',
            'iPhone 15 Pro 512GB', 'iPhone 15 Pro Max 256GB',
            'iPhone 16 128GB', 'iPhone 16 256GB',
            'iPhone 16 Pro 256GB', 'iPhone 16 Pro Max 256GB',
        ]:
            db.session.add(PhoneModel(name=name))
    db.session.commit()
    # Add demo data if DB is empty
    if not Phone.query.first():
        _add_demo_phones()

with app.app_context():
    init_db()

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
